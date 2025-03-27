import os
import re
import yaml
import json
import openpyxl as op
import operator


def write_excel(all_data,pass_data,fail_data,func_fail):
    # 写入excel，生成文件
    wb = op.Workbook() # 创建工作簿对象

    # sheet1: summary_results
    sr = wb.create_sheet("summary_results")
    sr.append(['all cases','passed cases','failed cases','func failed cases'])
    sr.append([len(all_data),len(pass_data),len(fail_data),len(func_fail)])
    # 功能失败的测例
    if len(func_fail) >= 1:
        # 写入总结果
        sr.append(['func fail case name',"results"])
        for key,value in func_fail.items():
            c = key,"func failed,need to check log"
            sr.append(c) # 每次写入一行
    else:
        pass
    # sheet2: summary_data
    sd = wb.create_sheet("summary_data")
    sd.append(['case_name','diff2','diff2_threshold','diff3_max','diff3_min','h2d_time(ms)','infer_time(ms)','d2h_time(ms)','tvm_e2e_time(ms)','ort_e2e_time(ms)','eval results']) # 添加表头
    # pass的测例
    if len(pass_data) >= 1:
        for key,value in pass_data.items():
            a = key,value[0],value[1],value[3],value[4],value[5],value[6],value[7],value[8],value[9],value[2]
            sd.append(a) # 每次写入一行
    else:
        pass

    # fail的测例
    if len(fail_data) >= 1:
        for key,value in fail_data.items():
            b = key,value[0],value[1],value[3],value[4],value[5],value[6],value[7],value[8],value[9],value[2]
            sd.append(b) #
    else:
        pass

    # 删除默认的sheet
    if wb["Sheet"]:
        del wb["Sheet"]
    else:
        pass

    path = "./logs"
    wb.save(os.path.join(path,"teco_infer_results.xlsx"))
    print("\n")
    print("- - - - - - - - - - - - - - - - - - - - - - - - -  - - - - - -  - - - - ")
    print(" teco_infer_results.xlsx is generated in \"./logs\" path !")
    print("- - - - - - - - - - - - - - - - - - - - - - - - -  - - - - - -  - - - - ")


def parse_log(log_path,infer_data_path,opts_status):

    if opts_status=="profiler":
        print("Now Open the Profiler Mode!")
    else:
        file_list = os.listdir(log_path)
        all_data = {}
        fail_data = {}
        func_fail = {}
        pass_data = {}
        log_yaml_f = infer_data_path + "/diff2_thresh.yaml"
        with open(log_yaml_f,'r',encoding='utf-8') as f:
            yaml_data = yaml.load(f.read(),Loader=yaml.FullLoader)
            # print(yaml_data)

        for i in file_list:
            case_path = log_path + i
            log_name = i
            with open (case_path,"r",encoding='UTF-8') as f:
                content = f.read()
                sw_eval = re.findall(r"diff2 = (\d+\.?\d*)",content)
                diff3_max = re.findall(r"diff3_max = (\d+\.?\d*)",content)
                diff3_min = re.findall(r"diff3_mean = (\d+\.?\d*)",content)
                h2d_time = re.findall(r"TecoInfer H2D time is: (\d+\.?\d*)",content) or re.findall(r"TVM H2D time is: (\d+\.?\d*)",content) or re.findall(r"TVM Includes codegen H2D time is: (\d+\.?\d*)",content)
                infer_time = re.findall(r"TecoInfer INFER time is: (\d+\.?\d*)",content) or re.findall(r"TVM INFER time is: (\d+\.?\d*)",content) or re.findall(r"TVM Includes codegen INFER time is: (\d+\.?\d*)",content)
                d2h_time = re.findall(r"TecoInfer D2H time is: (\d+\.?\d*)",content) or re.findall(r"TVM D2H time is: (\d+\.?\d*)",content) or re.findall(r"TVM Includes codegen D2H time is: (\d+\.?\d*)",content)
                e2e_time = re.findall(r"TecoInfer E2E time is: (\d+\.?\d*)",content) or re.findall(r"TVM E2E time is: (\d+\.?\d*)",content) or re.findall(r"TVM Includes codegen E2E time is: (\d+\.?\d*)",content)
                ort_time = re.findall(r"ORT E2E time is: (\d+\.?\d*)",content)

            # if "ppyolovel" in log_name:
            #     golden_data = 0.005
            # elif "rec" in log_name:
            #     golden_data = 0.01
            # elif "uie" in log_name or "det" in log_name:
            #     if "uie-bs1" in log_name or "uie-bs8" in log_name:
            #         golden_data = 0.06
            #     else:
            #         golden_data = 0.05
            # # resnet50, yolov5m，yolov5l6, yolov5x6, cls, unet, w2v
            # else:
            #     golden_data = 0.003

            for r in range(len(yaml_data)):
                base_name = yaml_data[r]["case_path"].split('/')[-1]
                if base_name == log_name:
                    golden_data = yaml_data[r]["diff2_threshold"]
                    break
                else:
                    golden_data = 0.1

            if sw_eval:
                # 有一个输出的测例
                if len(sw_eval) == 1:
                        if float(sw_eval[0]) <= float(golden_data):
                            save_data = [sw_eval[0],golden_data,"passed",diff3_max[0],diff3_min[0],h2d_time[0],infer_time[0],d2h_time[0],e2e_time[0],ort_time[0]]
                            pass_data.setdefault(log_name,save_data)
                            all_data.setdefault(log_name,save_data)
                        else:
                            save_data = [sw_eval[0],golden_data,"failed",diff3_max[0],diff3_min[0],h2d_time[0],infer_time[0],d2h_time[0],e2e_time[0],ort_time[0]]
                            fail_data.setdefault(log_name,save_data)
                            all_data.setdefault(log_name,save_data)
                # 有多个输出的测例
                elif len(sw_eval) > 1:
                    # golden_data比sw_eval大，返回True, 否则返会False.   ge即大于等于
                    eval_status = (all([operator.ge(float(golden_data), float(i)) for i in sw_eval]))
                    # print(eval_status)
                    multi_eval = ""
                    for d in sw_eval:
                        multi_eval =  str(d) + "," + multi_eval
                    # print(multi_eval)
                    if eval_status == True:
                        save_data = [multi_eval,golden_data,"passed",diff3_max[0],diff3_min[0],h2d_time[0],infer_time[0],d2h_time[0],e2e_time[0],ort_time[0]]
                        pass_data.setdefault(log_name,save_data)
                        all_data.setdefault(log_name,save_data)
                    else:
                        save_data = [multi_eval,golden_data,"failed",diff3_max[0],diff3_min[0],h2d_time[0],infer_time[0],d2h_time[0],e2e_time[0],ort_time[0]]
                        fail_data.setdefault(log_name,save_data)
                        all_data.setdefault(log_name,save_data)

            else:
                func_fail.setdefault(log_name,log_name)
                all_data.setdefault(log_name,log_name)

        print("**************************** summary teco infer results ****************************")

        if len(fail_data) == 0 and len(func_fail) == 0:
            print("ALL run %d cases, ALL passed! " % len(all_data))
            print("\n")
            print("* Passed Results:")
            for key,value in pass_data.items():
                print("\n")
                print("case_name:  %s, sw_eval:  %s  diff2:  %s diff3_max: %s  diff3_min: %s h2d_time(ms): %s infer_time(ms): %s d2h_time(ms): %s tvm_e2e_time(ms): %s ort_e2e_time(ms): %s eval results: %s" % (key,value[0],value[1],value[3],value[4],value[5],value[6],value[7],value[8],value[9],value[2]))
            write_excel(all_data,pass_data,fail_data,func_fail)
        else:
            print("--- all results ---")
            print("ALL run %d cases, %d cases passed, %d cases failed,  %d cases func failed !" % (len(all_data),len(pass_data),len(fail_data),len(func_fail)))

            print("\n")
            print("* func fail results:",len(func_fail))
            for key,value in func_fail.items():
                print("\n")
                print("case_name:  %s, func failed, need to check!" % (key))

            print("\n")
            print("* passed cases results:",len(pass_data))
            for key,value in pass_data.items():
                print("\n")
                print("case_name:  %s, diff2:  %s  diff2_threshold:  %s diff3_max: %s  diff3_min: %s h2d_time(ms): %s infer_time(ms): %s d2h_time(ms): %s tvm_e2e_time(ms): %s ort_e2e_time(ms): %s eval results: %s" % (key,value[0],value[1],value[3],value[4],value[5],value[6],value[7],value[8],value[9],value[2]))

            print("\n")
            print("* failed cases results:",len(fail_data))
            for key,value in fail_data.items():
                print("\n")
                print("case_name:  %s, diff2:  %s  diff2_threshold:  %s diff3_max: %s  diff3_min: %s h2d_time(ms): %s infer_time(ms): %s d2h_time(ms): %s tvm_e2e_time(ms): %s ort_e2e_time(ms): %s eval results: %s" % (key,value[0],value[1],value[3],value[4],value[5],value[6],value[7],value[8],value[9],value[2]))
            write_excel(all_data,pass_data,fail_data,func_fail)
            import sys
            sys.exit(1)


if __name__ == '__main__':
    log_path='./logs/'
    infer_data_path='./json/'
    parse_log(log_path,infer_data_path,"")
