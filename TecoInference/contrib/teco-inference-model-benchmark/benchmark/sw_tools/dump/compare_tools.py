# BSD 3- Clause License Copyright (c) 2023, Tecorigin Co., Ltd. All rights
# reserved.
# Redistribution and use in source and binary forms, with or without
# modification, are permitted provided that the following conditions are met:
# Redistributions of source code must retain the above copyright notice,
# this list of conditions and the following disclaimer.
# Redistributions in binary form must reproduce the above copyright notice,
# this list of conditions and the following disclaimer in the documentation
# and/or other materials provided with the distribution.
# Neither the name of the copyright holder nor the names of its contributors
# may be used to endorse or promote products derived from this software
# without specific prior written permission.
#
# THIS SOFTWARE IS PROVIDED BY THE COPYRIGHT HOLDERS AND CONTRIBUTORS "AS IS"
# AND ANY EXPRESS OR IMPLIED WARRANTIES, INCLUDING, BUT NOT LIMITED TO, THE
# IMPLIED WARRANTIES OF MERCHANTABILITY AND FITNESS FOR A PARTICULAR PURPOSE
# ARE DISCLAIMED. IN NO EVENT SHALL THE COPYRIGHT HOLDER OR CONTRIBUTORS BE
# LIABLE FOR ANY DIRECT, INDIRECT, INCIDENTAL, SPECIAL, EXEMPLARY, OR
# CONSEQUENTIAL DAMAGES (INCLUDING, BUT NOT LIMITED TO, PROCUREMENT OF
# SUBSTITUTE GOODS OR SERVICES; LOSS OF USE, DATA, OR PROFITS; OR BUSINESS
# INTERRUPTION)
# HOWEVER CAUSED AND ON ANY THEORY OF LIABILITY, WHETHER IN CONTRACT,
# STRICT LIABILITY,OR TORT (INCLUDING NEGLIGENCE OR OTHERWISE)  ARISING IN ANY
# WAY OUT OF THE USE OF THIS SOFTWARE, EVEN IF ADVISED OF THE POSSIBILITY
# OF SUCH DAMAGE.
# pylint: disable=bare-except
import os
import sys
import numpy as np
from utils.precision_standard import diff2, diff3_max, diff3_mean
import openpyxl as op_xl


class process_data(object):

    @staticmethod
    def cal_diff(eval_data, base_data, rtol=1e-7, atol=1e-7):
        # 均方相对误差开方 diff2
        diff2_result = diff2(eval_data, base_data)

        # 根据数量级选择相对误差或绝对误差 diff3
        try:
            np.testing.assert_allclose(eval_data, base_data, rtol=rtol, atol=atol, verbose=True)
            allclose = True
        except:
            allclose = False

        return diff2_result, allclose

    @staticmethod
    def diff3_result(sdaa_res, cpu_res, golden_res, th1):
        diff3_max_gold = diff3_max(cpu_res, golden_res, th1)
        diff3_max_sw = diff3_max(sdaa_res, golden_res, th1)

        diff3_mean_gold = diff3_mean(cpu_res, golden_res, th1)
        diff3_mean_sw = diff3_mean(sdaa_res, golden_res, th1)
        return (diff3_max_sw, diff3_max_gold, diff3_mean_sw, diff3_mean_gold)


class Diff3CompareData(object):

    def __init__(self, sdaa_dir, ort_dir, golden_dir, output_dir, input_tag, th1, th2):
        self.sdaa_dir = sdaa_dir
        self.ort_dir = ort_dir
        self.golden_dir = golden_dir
        self.output_dir = output_dir
        self.input_tag = input_tag
        self.th1 = th1
        self.th2 = th2
        if self.golden_dir != "":
            if not os.path.isdir(self.sdaa_dir) or not os.path.isdir(
                    self.ort_dir) or not os.path.isdir(self.golden_dir):
                raise ValueError(
                    print("analysis_tools.py -s {} -d {} -g {} should be directory.".format(
                        self.sdaa_dir, self.ort_dir, self.golden_dir)))
        else:
            if not os.path.isdir(self.sdaa_dir) or not os.path.isdir(self.ort_dir):
                raise ValueError(
                    print("analysis_tools.py -s {} -d {} should be directory.".format(
                        self.sdaa_dir, self.ort_dir)))

    def check_abnormal(self, file_path, actual) -> bool:
        if np.any(np.isnan(actual)) | np.any(np.isinf(actual)):
            print("file path : {}, actual val: {}, has nan: {}, has inf: {}.".format(
                file_path, actual, np.any(np.isnan(actual)), np.any(np.isinf(actual))))
            return True
        return False

    def compare_three_folders(self, folder1, folder2, folder3):
        # 获取两个文件夹中的文件列表
        files1 = set(os.listdir(folder1))
        files2 = set(os.listdir(folder2))
        # 找出在三个文件夹中都存在的文件 并对其进行字典序排序
        common_files = files1 & files2
        sort_common_files = sorted(common_files)
        # 比较在三个文件夹中都存在的文件的内容
        if folder3 != "":
            files3 = set(os.listdir(folder3))
            common_files = common_files & files3
            sort_common_files = sorted(common_files)
        order_index = 0
        input_tag = self.input_tag
        dic_list = {}
        for file in sort_common_files:
            file1 = os.path.join(folder1, file)
            file2 = os.path.join(folder2, file)
            file1_npy_array = np.load(file1)
            file2_npy_array = np.load(file2)
            check_val = self.check_abnormal(file1, file1_npy_array) | self.check_abnormal(
                file2, file2_npy_array)
            if np.shape(file1_npy_array) != np.shape(file2_npy_array):
                print(
                    "FAILED: [file1: {}] vs [file2: {}] " \
                        "shape mismatch, shape1: {} vs shape2: {}"
                    .format(file1, file2, np.shape(file1_npy_array),
                            np.shape(file2_npy_array)))
                with open(self.output_dir, "w") as fd:
                    fd.write(
                        "[file1: {}] vs [file2: {}] shape mismatch," \
                        " shape1: {} vs shape2: {}. check input file shape! "
                        .format(file1, file2, np.shape(file1_npy_array),
                                np.shape(file2_npy_array)))
                sys.exit()
            actual = np.asanyarray(file1_npy_array).astype("float32")
            desired = np.asanyarray(file2_npy_array).astype("float32")
            if folder3 != "":
                file3 = os.path.join(folder3, file)
                file3_npy_array = np.load(file3)
                check_val = check_val | self.check_abnormal(file3, file3_npy_array)
                if np.shape(file1_npy_array) != np.shape(file3_npy_array):
                    print(
                        "FAILED: [file1: {}] vs [file2: {}] shape mismatch, " \
                        "shape1: {} vs shape2: {}"
                        .format(file1, file3, np.shape(file1_npy_array),
                                np.shape(file3_npy_array)))
                    with open(self.output_dir, "w") as fd:
                        fd.write(
                            "[file1: {}] vs [file2: {}] shape mismatch, " \
                            "shape1: {} vs shape2: {}. check input file shape! "
                            .format(file1, file3, np.shape(file1_npy_array),
                                    np.shape(file3_npy_array)))
                    sys.exit()
                diff3_max_sw, diff3_max_gold, diff3_mean_sw, \
                diff3_mean_gold = process_data.diff3_result(
                    file1_npy_array, file2_npy_array, file3_npy_array,
                    [self.th1])
                diff3_max_val = "pass" if diff3_max_sw <= diff3_max_gold * self.input_tag \
                                        else "fail"
                diff3_mean_val = "pass" if diff3_mean_sw <= diff3_mean_gold * self.input_tag \
                                        else "fail"
                diff2_val = diff2(actual, desired)
                tmp_list = [("file", file), ("diff2_val", diff2_val),
                            ("diff3_max_val", diff3_max_val), ("diff3_mean_val", diff3_mean_val),
                            ("diff3_max_sw", diff3_max_sw), ("diff3_max_gold", diff3_max_gold),
                            ("diff3_mean_sw", diff3_mean_sw), ("diff3_mean_gold", diff3_mean_gold),
                            ("input_tag", input_tag), ("check_val", check_val)]
                tmp_dic_list = dict(tmp_list)
                dic_list.setdefault(order_index, tmp_dic_list)
            else:
                diff2_val, allclose = process_data.cal_diff(actual, desired, self.th1, self.th1)
                tmp_list = [("file", file), ("check_val", check_val), ("diff2_val", diff2_val),
                            ("allclose", allclose)]
                tmp_dic_list = dict(tmp_list)
                dic_list.setdefault(order_index, tmp_dic_list)
            order_index = order_index + 1
        return dic_list

    def process_three_dic(self, dic_list, file_path):
        if os.path.exists(file_path):
            os.remove(file_path)
        with open(file_path, "w") as file:
            for key, dic_item in dic_list.items():
                if len(dic_item.values()) > 5:
                    file.write("index: {} | {} | nan_inf: {} | diff2: {} | " \
                        "diff3_max={} | diff3_mean={} | " \
                        "diff3_max_sw: {} | "\
                        "diff3_max_g: {}*{}={} | " \
                        "diff3_mean_sw: {} | " \
                        "diff3_mean_g: {}*{}={}".
                        format(key,
                                dic_item["file"],
                                dic_item["check_val"],
                                dic_item["diff2_val"],
                                dic_item["diff3_max_val"],
                                dic_item["diff3_mean_val"],
                                dic_item["diff3_max_sw"],
                                dic_item["diff3_max_gold"],
                                dic_item["input_tag"],
                                dic_item["diff3_max_gold"] * dic_item["input_tag"],
                                dic_item["diff3_mean_sw"],
                                dic_item["diff3_mean_gold"],
                                dic_item["input_tag"],
                                dic_item["diff3_mean_gold"] * dic_item["input_tag"]))
                    file.write("\n")
                else:
                    file.write("index: {} | {} | nan_inf: {} | "\
                               "diff2: {} | all_close: {} | th: {}"
                               .format(key, dic_item["file"], dic_item["check_val"],
                                        dic_item["diff2_val"], dic_item["allclose"], self.th1))
                    file.write("\n")

    def folders_msg(self, folder1, folder2, folder3, file_path):
        files1 = set(os.listdir(folder1))
        files2 = set(os.listdir(folder2))
        only_in_folder2 = files2 - files1
        common_files = files1 & files2
        only_in_folder1 = files1 - files2
        with open(file_path, "a") as file:
            if folder3 != "":
                files3 = set(os.listdir(folder3))
                common_files = common_files & files3
                only_in_folder3 = files3 - files1
                file.write("Files only in {}: {}".format(folder3, only_in_folder3))
                file.write("\n")
            # 找出只在其中一个文件夹中的文件
            file.write("Common files with differences: {}".format(common_files))
            file.write("\n")
            file.write("Files only in {}: {}".format(folder1, only_in_folder1))
            file.write("\n")
            file.write("Files only in {}: {}".format(folder2, only_in_folder2))
            file.write("\n")

    def excel_msg(self, dic_list, folder, write_path):
        nan_inf_list = []
        diff2_list = []
        diff3_list = []
        file_path = "compare_results.xlsx"
        if os.path.exists(file_path):
            os.remove(file_path)
        wb = op_xl.Workbook()
        ws = wb['Sheet']
        ws2 = wb.create_sheet(title='nan_inf')
        ws3 = wb.create_sheet(title='diff2_val')
        ws3.append(["file", "diff2_val", "threshold"])
        if "" != folder:
            ws4 = wb.create_sheet(title='diff3_val')
            ws4.append(
                ["file", "diff3_max_sw", "diff3_max_gold", "diff3_mean_sw", "diff3_mean_gold"])
            ws.append([
                "index", "file", "nan_inf", "diff2_val", "diff3_max_val", "diff3_mean_val",
                "diff3_max_sw", "diff3_max_gold", "diff3_mean_sw", "diff3_mean_gold", "input_tag"
            ])
            for key, dic_item in dic_list.items():
                ws.append([
                    key, dic_item["file"], dic_item["check_val"], dic_item["diff2_val"],
                    dic_item["diff3_max_val"], dic_item["diff3_mean_val"], dic_item["diff3_max_sw"],
                    dic_item["diff3_max_gold"] * dic_item["input_tag"], dic_item["diff3_mean_sw"],
                    dic_item["diff3_mean_gold"] * dic_item["input_tag"], dic_item["input_tag"]
                ])
                if True is dic_item["check_val"]:
                    ws2.append([dic_item["file"]])
                    nan_inf_list.append(dic_item["file"])
                if self.th2 < dic_item["diff2_val"]:
                    ws3.append([dic_item["file"], dic_item["diff2_val"], self.th2])
                    diff2_list.append(dic_item["file"])
                if dic_item["diff3_max_val"] != "pass" or dic_item["diff3_mean_val"] != "pass":
                    ws4.append([
                        dic_item["file"], dic_item["diff3_max_sw"],
                        dic_item["diff3_max_gold"] * dic_item["input_tag"],
                        dic_item["diff3_mean_sw"],
                        dic_item["diff3_mean_gold"] * dic_item["input_tag"]
                    ])
                    diff3_list.append(dic_item["file"])

        else:
            ws4 = wb.create_sheet(title='all_close')
            ws.append(["index", "file", "nan_inf", "diff2", "all_close", "th"])
            for key, dic_item in dic_list.items():
                ws.append([
                    key, dic_item["file"], dic_item["check_val"], dic_item["diff2_val"],
                    dic_item["allclose"], self.th1
                ])
                if True is dic_item["check_val"]:
                    ws2.append([dic_item["file"]])
                    nan_inf_list.append(dic_item["file"])
                if self.th2 < dic_item["diff2_val"]:
                    ws3.append([dic_item["file"], dic_item["diff2_val"], self.th2])
                    diff2_list.append(dic_item["file"])
                if True is not dic_item["allclose"]:
                    ws4.append([dic_item["file"]])
                    diff3_list.append(dic_item["file"])
        wb.save(os.path.join("compare_results.xlsx"))
        with open(write_path, "a") as fd:
            fd.write("------------------------- summary message -------------------------\n")
            fd.write("nan_inf list: {}".format(nan_inf_list))
            fd.write("\n")
            fd.write("------------------------- diff2 list message -------------------------\n")
            fd.write("diff2 list: {}".format(diff2_list))
            fd.write("\n")
            fd.write("------------------------- diff3 list message -------------------------\n")
            fd.write("other list: {}".format(diff3_list))
            fd.write("\n")

    def compare_data(self):
        file_dic = self.compare_three_folders(self.sdaa_dir, self.ort_dir, self.golden_dir)
        self.process_three_dic(file_dic, self.output_dir)
        self.folders_msg(self.sdaa_dir, self.ort_dir, self.golden_dir, self.output_dir)
        self.excel_msg(file_dic, self.golden_dir, self.output_dir)
        print("success compare sdaa vs ort tensor data !")


class Diff3Process(object):

    def __init__(self, sdaa_dir, ort_dir, golden_dir, output_dir, input_tag, th1, th2):
        self.diff3_cmp = Diff3CompareData(sdaa_dir, ort_dir, golden_dir, output_dir, input_tag, th1,
                                          th2)

    def process_data(self):
        self.diff3_cmp.compare_data()


class CompareDiff3Handle(object):

    def __init__(self,
                 sdaa_dir,
                 ort_dir,
                 golden_dir,
                 output_dir,
                 input_tag=10,
                 th1=0.000001,
                 th2=0.001):
        self.sdaa_dir = sdaa_dir
        self.ort_dir = ort_dir
        self.golden_dir = golden_dir
        self.output_dir = output_dir
        self.input_tag = input_tag
        self.th1 = th1
        self.th2 = th2
        self.framework_process = Diff3Process(self.sdaa_dir, self.ort_dir, self.golden_dir,
                                              self.output_dir, self.input_tag, self.th1, self.th2)

    def operation(self):
        self.framework_process.process_data()
